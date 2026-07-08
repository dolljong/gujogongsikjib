# -*- coding: utf-8 -*-
"""
구조공식집 엑셀 -> JSON 변환기.

7개 xlsx를 일괄 처리한다. 열을 헤더 키워드로 자동 감지하므로
파일마다 열 구성이 조금 달라도 동작한다.

출력: data/formulas.json  (케이스 배열)
케이스 = 같은 NO(공식번호)로 묶은 행 그룹.
프론트엔드가 category/subCategory/classes 로 트리를 구성한다.
"""
import openpyxl, glob, os, io, json, re, shutil

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_DIR = os.path.join(ROOT, "공식엑셀")
FIG_DIR = os.path.join(ROOT, "삽도")
OUT = os.path.join(ROOT, "app", "src", "data", "formulas.json")
FIG_OUT = os.path.join(ROOT, "app", "public", "figures")

IMG_EXTS = [".png", ".jpg", ".jpeg", ".gif", ".svg"]

# 원본 엑셀 LaTeX 오타 보정. {케이스 id: [(찾을 문자열, 바꿀 문자열)]}
# 계산기가 원본 공식을 그대로 쓰므로, 명백한 오타는 여기서 교정한다.
LATEX_FIXES = {
    # 중공 직사각형 I식: 아래첨자 오타 h_i → h_1 (Z·i 식은 h_1 로 정상)
    "1.6.11": [("h_{i}", "h_{1}")],
}


def apply_latex_fix(cid, latex):
    if latex is None:
        return latex
    for old, new in LATEX_FIXES.get(str(cid), []):
        latex = latex.replace(old, new)
    return latex


def build_figure_folders():
    """삽도 하위 폴더를 앞 번호(0,1,2..)로 매핑. 엑셀 파일명과 폴더명의
    공백/표기 차이를 흡수한다."""
    folders = {}
    if not os.path.isdir(FIG_DIR):
        return folders
    for name in os.listdir(FIG_DIR):
        p = os.path.join(FIG_DIR, name)
        if os.path.isdir(p):
            m = re.match(r"^\s*(\d+)", name)
            if m:
                folders[int(m.group(1))] = p
    return folders


FIG_FOLDERS = build_figure_folders()


def detect_columns(header):
    """헤더 행에서 열 역할을 감지. {role: col_index(1-base)} + classes[col_index]."""
    roles = {}
    class_cols = []
    for idx, h in enumerate(header, start=1):
        if h is None:
            continue
        s = str(h).strip()
        low = s.lower()
        if "page" in low:
            roles["page"] = idx
        elif low in ("no", "case") or low == "no.":
            roles["id"] = idx
        elif "latex" in low:
            roles["latex"] = idx
        elif "변수" in s:
            roles["variables"] = idx
        elif "이름" in s:                      # 공식 이름 / case 이름 = 결과명
            roles["name"] = idx
        elif "설명" in s:                      # 공식설명(재료역학) = 케이스 설명
            roles["desc"] = idx
        else:
            class_cols.append((idx, s))        # 형태 / 하중형태 / 하중형식 ...
    return roles, class_cols


def cell(ws, r, c):
    if c is None:
        return None
    v = ws.cell(r, c).value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _copy_fig(src):
    os.makedirs(FIG_OUT, exist_ok=True)
    name = os.path.basename(src)
    shutil.copyfile(src, os.path.join(FIG_OUT, name))
    return f"figures/{name}"


def find_figures(fig_folder, case_id):
    """삽도 폴더에서 case_id 이미지를 모두 찾아 복사하고 상대경로 리스트를 반환.
    - 정확일치 `<id>.png` 가 있으면 그 한 장.
    - 없으면 `<id>.N.png`(N=숫자) 여러 장을 번호순으로 (계수표·형상군 삽도)."""
    if not fig_folder or not os.path.isdir(fig_folder):
        return []
    for ext in IMG_EXTS:                       # 1) 정확일치 우선
        src = os.path.join(fig_folder, f"{case_id}{ext}")
        if os.path.exists(src):
            return [_copy_fig(src)]
    # 2) 다중 삽도: <id>.<숫자>.<ext>
    multi = []
    pat = re.compile(r"^" + re.escape(str(case_id)) + r"\.(\d+)\.(" +
                     "|".join(e.lstrip(".") for e in IMG_EXTS) + r")$", re.I)
    for name in os.listdir(fig_folder):
        m = pat.match(name)
        if m:
            multi.append((int(m.group(1)), os.path.join(fig_folder, name)))
    multi.sort(key=lambda t: t[0])
    return [_copy_fig(p) for _, p in multi]


def category_meta(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    m = re.match(r"^\s*(\d+)[.\s]+(.*)$", base)
    order = int(m.group(1)) if m else 999
    name = (m.group(2) if m else base).strip()
    return order, name, base


def main():
    cases = []
    for f in sorted(glob.glob(os.path.join(XLSX_DIR, "*.xlsx"))):
        order, cat_name, base = category_meta(f)
        fig_folder = FIG_FOLDERS.get(order)
        wb = openpyxl.load_workbook(f, data_only=False)
        multi = len(wb.worksheets) > 1

        for ws in wb.worksheets:
            header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            roles, class_cols = detect_columns(header)
            sub = ws.title.strip() if multi else None

            # 케이스 그룹핑용 누적 딕셔너리 (파일+시트 범위 내)
            grouped = {}
            order_seen = []
            auto = 0
            for r in range(2, ws.max_row + 1):
                latex = cell(ws, r, roles.get("latex"))
                name = cell(ws, r, roles.get("name"))
                variables = cell(ws, r, roles.get("variables"))
                desc = cell(ws, r, roles.get("desc"))
                classes = [cell(ws, r, ci) for ci, _ in class_cols]
                page = cell(ws, r, roles.get("page"))
                rid = cell(ws, r, roles.get("id"))

                # 완전 빈 행 스킵
                if latex is None and name is None and not any(classes) and desc is None:
                    continue

                # 케이스 id 결정
                if rid is not None:
                    cid = str(rid).strip()
                else:
                    auto += 1
                    cid = f"{base}#{auto}"    # NO 없는 파일(재료역학 등) = 행마다 케이스

                key = (sub, cid)
                if key not in grouped:
                    grouped[key] = {
                        "category": cat_name,
                        "categoryOrder": order,
                        "subCategory": sub,
                        "id": cid,
                        "classes": [c for c in classes if c],
                        "page": page,
                        "figures": find_figures(fig_folder, cid),
                        "desc": desc,
                        "results": [],
                    }
                    order_seen.append(key)
                g = grouped[key]
                # 첫 행에서 못 채운 메타 보강
                if not g["figures"]:
                    g["figures"] = find_figures(fig_folder, cid)
                if not g["classes"]:
                    g["classes"] = [c for c in classes if c]
                if g["page"] is None:
                    g["page"] = page
                if g["desc"] is None:
                    g["desc"] = desc

                g["results"].append({
                    "name": name,
                    "latex": apply_latex_fix(cid, latex),
                    "variables": variables,
                    # Phase 2(계산기)에서 채울 필드:
                    "expr": None,
                    "outputs": [],
                })

            for key in order_seen:
                cases.append(grouped[key])

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with io.open(OUT, "w", encoding="utf-8") as fp:
        json.dump(cases, fp, ensure_ascii=False, indent=1)

    # 요약
    total_results = sum(len(c["results"]) for c in cases)
    with_fig = sum(1 for c in cases if c["figures"])
    print(f"cases        : {len(cases)}")
    print(f"result rows  : {total_results}")
    print(f"with figure  : {with_fig}/{len(cases)}")
    by_cat = {}
    for c in cases:
        by_cat.setdefault(c["category"], 0)
        by_cat[c["category"]] += 1
    for k, v in by_cat.items():
        print(f"  - {k}: {v} cases")
    print(f"-> {OUT}")


if __name__ == "__main__":
    main()
